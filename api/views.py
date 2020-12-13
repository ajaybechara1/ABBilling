from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view
from .serializers import CustomerCompanyDetailSerializer, TruckDetailSerializer, CityDetailSerializer, PackageDetailSerializer, OrderTypeSerializer, BillDetailSerializer
from .models import CustomerCompanyDetail, TruckDetail, CityDetail, PackageDetail, OrderType, BillDetail, DummyModel

from django.views.decorators.csrf import csrf_exempt
import sys
import datetime

import time

import xlwt
from xlwt import Workbook
from openpyxl import Workbook as Openpyxl_Workbook
from openpyxl import load_workbook as Openpyxl_load_workbook
from openpyxl.styles import Font as Openpyxl_Style
from num2words import num2words

global_dictionary = "None"

class CustomerCompanyDetailView(viewsets.ModelViewSet):
	serializer_class = CustomerCompanyDetailSerializer
	queryset = CustomerCompanyDetail.objects.all()

class TruckDetailView(viewsets.ModelViewSet):
	serializer_class = TruckDetailSerializer
	queryset = TruckDetail.objects.all()

class CityDetailView(viewsets.ModelViewSet):
	serializer_class = CityDetailSerializer
	queryset = CityDetail.objects.all()

class PackageDetailView(viewsets.ModelViewSet):
	serializer_class = PackageDetailSerializer
	queryset = PackageDetail.objects.all()

class OrderTypeView(viewsets.ModelViewSet):
	serializer_class = OrderTypeSerializer
	queryset = OrderType.objects.all()

class BillDetailView(viewsets.ModelViewSet):
	serializer_class = BillDetailSerializer
	queryset = BillDetail.objects.all()

@api_view(['POST'])
def AddNewOrder(request):
	try:
		data = request.data

		print('\n\n\n\n', data, '\n\n\n\n')

		pk = data['company_name'][:-1].split('/')[-1]
		company_name = CustomerCompanyDetail.objects.get(id=int(pk))

		pk = data['truck_number'][:-1].split('/')[-1]
		truck_number = TruckDetail.objects.get(id=pk)

		pk = data['package_size'][:-1].split('/')[-1]
		package_size = PackageDetail.objects.get(id=pk)

		pk = data['city_from'][:-1].split('/')[-1]
		city_from = CityDetail.objects.get(id=pk)

		pk = data['city_to'][:-1].split('/')[-1]
		city_to = CityDetail.objects.get(id=pk)

		pk = data['order_type'][:-1].split('/')[-1]
		order_type = OrderType.objects.get(id=pk)


		BillDetail.objects.create(
				order_date = data['order_date'],
				company_name = company_name,
				truck_number = truck_number,
				package_size = package_size,
				city_from = city_from,
				city_to = city_to,
				order_type = order_type,
				container_number = data['container_number'],
				lr_number = data['lr_number'],
				freight = data['freight'],
				overload = data['overload'],
				liftoff = data['liftoff'],
				advanced = data['advanced'],
				diesel = data['diesel'],
				balance = data['balance'],
			)
		return Response({}, status=status.HTTP_201_CREATED)
	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
		print(str(e), str(exc_tb.tb_lineno))
		return Response({}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def TakeBackup(request):
	try:
		import shutil

		original = r'db.sqlite3'

		d = datetime.datetime.now()
		file_name = d.strftime("%d-%B-%Y_%I:%M_%p_%A") + "_db"
		file_name = "files/backup/" + file_name + ".sqlite3"

		shutil.copyfile(original, file_name)

		return Response({}, status=status.HTTP_201_CREATED)
	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
		print(str(e), str(exc_tb.tb_lineno))
		return Response({}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def AddNewOrder2(request):
	data = request.data
	serializer = BillDetailSerializer(data=data)
	if serializer.is_valid():
		print(serializer.validated_data)
		return Response(serializer.data, status=status.HTTP_201_CREATED)
	return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def get_values_for_filter(data):
	company_list = []
	for company_url in data['company_name']:
		pk = company_url[:-1].split('/')[-1]
		company_obj = CustomerCompanyDetail.objects.get(id=int(pk))
		company_list.append( company_obj )

	if not company_list:
		company_list = list(CustomerCompanyDetail.objects.all())

	truck_list = []
	for truck_url in data['truck_number']:
		pk = truck_url[:-1].split('/')[-1]
		truck_obj = TruckDetail.objects.get(id=int(pk))
		truck_list.append( truck_obj )

	if not truck_list:
		truck_list = list(TruckDetail.objects.all())

	package_list = []
	for package_url in data["package_size"]:
		pk = package_url[:-1].split('/')[-1]
		package_obj = PackageDetail.objects.get(id=int(pk))
		package_list.append( package_obj )

	if not package_list:
		package_list = list(PackageDetail.objects.all())

	city_from_list = []
	for city_from_url in data["city_from"]:
		pk = city_from_url[:-1].split('/')[-1]
		city_from_obj = CityDetail.objects.get(id=int(pk))
		city_from_list.append( city_from_obj )

	if not city_from_list:
		city_from_list = list(CityDetail.objects.all())

	city_to_list = []
	for city_to_url in data["city_to"]:
		pk = city_to_url[:-1].split('/')[-1]
		city_to_obj = CityDetail.objects.get(id=int(pk))
		city_to_list.append( city_to_obj )

	if not city_to_list:
		city_to_list = list(CityDetail.objects.all())

	order_list = []
	for order_url in data["order_type"]:
		pk = order_url[:-1].split('/')[-1]
		order_obj = OrderType.objects.get(id=int(pk))
		order_list.append( order_obj )

	if not order_list:
		order_list = list(OrderType.objects.all())

	return company_list, truck_list, package_list, city_from_list, city_to_list, order_list


def get_filtered_order_list(data):
	company_list, truck_list, package_list, city_from_list, city_to_list, order_list = get_values_for_filter(data)

	date_from = data["date_from"]
	date_to = data["date_to"]

	date_format = "%Y-%m-%d"
	datetime_start = datetime.datetime.strptime(
		date_from, date_format).date()
	datetime_end = datetime.datetime.strptime(date_to, date_format).date()  # noqa: F841

	order_list = BillDetail.objects.filter(
			order_date__gte=datetime_start,
			order_date__lte=datetime_end,
			company_name__in=company_list,
			truck_number__in=truck_list,
			package_size__in=package_list,
			city_from__in=city_from_list,
			city_to__in=city_to_list,
			order_type__in=order_list
		)

	order_list = order_list.order_by("order_date")
	order_list = list(order_list)

	response_order_list = []

	total_amount = 0
	total_liftoff = 0
	total_advanced = 0
	total_net_amount = 0
	total_freight = 0

	for order_obj in order_list:
		res_obj = {}
		res_obj['id'] = order_obj.id
		res_obj['company_name'] = order_obj.company_name.company_name
		res_obj['truck_number'] = order_obj.truck_number.truck_number
		res_obj['package_size'] = order_obj.package_size.package_size
		res_obj['city_from'] = order_obj.city_from.city_name
		res_obj['city_to'] = order_obj.city_to.city_name
		res_obj['order_type'] = order_obj.order_type.order_type
		res_obj['container_number'] = order_obj.container_number
		res_obj['lr_number'] = order_obj.lr_number
		res_obj['freight'] = order_obj.freight
		res_obj['overload'] = order_obj.overload
		res_obj['liftoff'] = order_obj.liftoff
		res_obj['advanced'] = order_obj.advanced
		res_obj['diesel'] = order_obj.diesel
		res_obj['tds'] = order_obj.tds
		res_obj['balance'] = order_obj.balance
		res_obj['order_date'] = str(order_obj.order_date)

		total_liftoff += float(order_obj.liftoff)
		total_advanced += float(order_obj.advanced)
		total_freight += float(order_obj.freight)

		response_order_list.append(res_obj)

	total_amount = total_freight - total_advanced
	total_net_amount = total_amount + total_liftoff

	return response_order_list, total_amount, total_liftoff, total_advanced, total_net_amount, total_freight


@api_view(['POST'])
def FilterOrder(request):
	response = {
		"status" : 500
	}
	try:
		data = request.data
		order_list, total_amount, total_liftoff, total_advanced, total_net_amount, total_freight = get_filtered_order_list(data)

		total_count = len(order_list)
		response = {
			"status" : 200,
			"order_list": order_list,
			"total_amount" : total_amount,
			"total_liftoff" : total_liftoff,
			"total_advanced" : total_advanced,
			"total_net_amount" : total_net_amount,
			"total_freight" : total_freight,
			"total_count": total_count,
		}

		return Response(response, status=status.HTTP_201_CREATED)
	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
		return Response({}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def GenerateBill(request):
	response = {
		"status" : 500
	}
	try:
		data = request.data
		order_list, total_amount, total_liftoff, total_advanced, total_net_amount, total_freight = get_filtered_order_list(data)
		bill_date = data["bill_date"]
		yyyy, mm, dd = bill_date.split("-")

		party_gst = data["party_gst"]

		party_name = order_list[0]["company_name"]
		excel_link = create_bill_with_excel_template(order_list, bill_number="GR/2020-21/", bill_date=dd+"/"+mm+"/"+yyyy, pan_number="AUXPK 3642 D", party_gst=party_gst, party_name=party_name)

		response = {
			"status" : 200,
			"link": excel_link
		}

		return Response(response, status=status.HTTP_201_CREATED)
	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
		print(e)
		return Response(response, status=status.HTTP_400_BAD_REQUEST)



# 737

@api_view(['GET'])
def dummy_request(request):

	global global_dictionary
	global_dictionary = "AJAY"
	# bill_obj = BillDetail.objects.all()[0]
	# start = time.time()

	# dummy_model_objs = bill_obj.dummy_model.all()
	# print('\n\n\n', dummy_model_objs)

	# end = time.time()

	# total = str(end - start)

	# start = time.time()

	# for i in range(1,737):
	# 	dummy_model_objs = DummyModel.objects.get(pk=i)

	# end = time.time()

	# total += ( "  " +  str( end-start ) )

	return Response(global_dictionary)


@api_view(['GET'])
def dummy_request2(request):

	global global_dictionary
	# global_dictionary = "AJAY"
	# bill_obj = BillDetail.objects.all()[0]
	# start = time.time()

	# dummy_model_objs = bill_obj.dummy_model.all()
	# print('\n\n\n', dummy_model_objs)

	# end = time.time()

	# total = str(end - start)

	# start = time.time()

	# for i in range(1,737):
	# 	dummy_model_objs = DummyModel.objects.get(pk=i)

	# end = time.time()

	# total += ( "  " +  str( end-start ) )

	return Response(global_dictionary)


def add_new_sheet_to_bill(sheet, orders):

	try:
		sheet.write_merge(0, 0, 0, 14, 	"Goodwill Roadways", 																xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(1, 1, 0, 14, 		"Suppliers of Trailors & Transport Contractor", 									xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(2, 2, 0, 14, 		"Plot no 101. Survey No 165 Meghpar Borichi Anjau Kutch     M:-8511331003", 		xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(3, 3, 0, 3,			"To : " +  "company_name" , 														xlwt.easyxf('align: horz center, vert center;'))

		sheet.write_merge(3, 3, 10, 10, "Bill NO : ", 																			xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(3, 3, 11, 14, "AB/2020/11/1", 																			xlwt.easyxf('align: horz center, vert center;'))

		sheet.write_merge(4, 4, 10, 10, "Date : ", 																				xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(4, 4, 11, 14, "2000/03/09", 																			xlwt.easyxf('align: horz center, vert center;'))

		sheet.write_merge(5, 5, 10, 10, "Duration : ", 																			xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(5, 5, 11, 14, "", 																						xlwt.easyxf('align: horz center, vert center;'))

		sheet.write_merge(6, 6, 10, 10, "PAN NO : ", 																				xlwt.easyxf('align: horz center, vert center;'))
		sheet.write_merge(6, 6, 11, 14, "AB344KAL", 																				xlwt.easyxf('align: horz center, vert center;'))

		sheet.write_merge(7, 7, 0, 14,  "Being the amount of freight charges debited to your account as per details shown below.", xlwt.easyxf('align: horz center, vert center;'))

		sheet.write(8 , 0 , "Date")
		sheet.write(8 , 1 , "L/R No.")
		sheet.write(8 , 2 , "Truck No.")
		sheet.write(8 , 3 , "Container No.")
		sheet.write(8 , 4 , "From")
		sheet.write(8 , 5 , "To")
		sheet.write(8 , 6 , "Pkgs.")
		sheet.write(8 , 7 , "Weight")
		sheet.write(8 , 8 , "Fright")
		sheet.write(8 , 9 , "Oveload")
		sheet.write(8 , 10 , "Liftoff")
		sheet.write(8 , 11, "Advanced")
		sheet.write(8 , 12, "Diesel")
		sheet.write(8 , 13 , "tds")
		sheet.write(8 , 14, "balance")

		total_freight = 0
		total_overload = 0
		total_Liftoff = 0
		total_advanced = 0
		total_diesel = 0
		total_tds = 0
		total_balance = 0

		row_number = 9
		for order in orders:
			print(order)
			total_freight += float(order['freight'])
			total_overload += float(order['overload'])
			total_Liftoff += float(order['liftoff'])
			total_advanced += float(order['advanced'])
			total_diesel += float(order['diesel'])
			total_tds += float(order['tds'])
			total_balance += float(order['balance'])

			sheet.write(row_number , 0 , order["order_date"])
			sheet.write(row_number , 1 , order["lr_number"])
			sheet.write(row_number , 2 , order["truck_number"])
			sheet.write(row_number , 3 , order["container_number"])
			sheet.write(row_number , 4 , order["city_from"])
			sheet.write(row_number , 5 , order["city_to"])
			sheet.write(row_number , 6 , order["package_size"])
			sheet.write(row_number , 7 , order["order_type"])
			sheet.write(row_number , 8 , order["freight"])
			sheet.write(row_number , 9 , order["overload"])
			sheet.write(row_number , 10 ,order["liftoff"])
			sheet.write(row_number , 11, order["advanced"])
			sheet.write(row_number , 12, order["diesel"])
			sheet.write(row_number , 13 ,order["tds"])
			sheet.write(row_number , 14, order["balance"])

			row_number += 1

		sheet.write_merge(row_number, row_number, 0, 0, "Total : ")

		sheet.write_merge(row_number, row_number, 8, 8, total_freight)
		sheet.write_merge(row_number, row_number, 9, 9, total_overload)
		sheet.write_merge(row_number, row_number, 10, 10, total_Liftoff)
		sheet.write_merge(row_number, row_number, 11, 11, total_advanced)
		sheet.write_merge(row_number, row_number, 12, 12, total_diesel)
		sheet.write_merge(row_number, row_number, 13, 13, total_tds)
		sheet.write_merge(row_number, row_number, 14, 14, total_balance)

	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
		print(str(e), str(exc_tb.tb_lineno))

def create_bill_with_excel(orders, bill_number="", bill_date="", pan_number=""):
    # file_path = "chat/mis_dashboard"
    try:

        work_book = Workbook()

        sheet = work_book.add_sheet("Bill1")
        add_new_sheet_to_bill(sheet, orders)

        sheet = work_book.add_sheet("Bill2")
        add_new_sheet_to_bill(sheet, orders)

        filename = "filtered_data_" + str(datetime.datetime.today())  + ".xls"

        work_book.save(filename)

        print("BHAI BHAI")
        return filename
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        print("OH NO")
        print(str(e), str(exc_tb.tb_lineno))


def load_data_into_sheet(orders, sheet):
	total_freight = 0
	total_overload = 0
	total_Liftoff = 0
	total_advanced = 0
	total_diesel = 0
	total_tds = 0
	total_balance = 0

	row_number = 9
	for order in orders:
		total_freight += float(order['freight'])
		total_Liftoff += float(order['liftoff'])
		total_advanced += float(order['advanced'])
		total_balance += (float(order["freight"]) - float(order["advanced"]))
		# total_overload += float(order['overload'])
		# total_diesel += float(order['diesel'])
		# total_tds += float(order['tds'])

		sheet['A' + str(row_number)] = order["order_date"]
		sheet['B' + str(row_number)] = order["truck_number"]
		sheet['C' + str(row_number)] = order["container_number"]
		sheet['D' + str(row_number)] = order["city_from"]
		sheet['E' + str(row_number)] = order["city_to"]
		sheet['F' + str(row_number)] = order["package_size"]
		sheet['G' + str(row_number)] = order["order_type"]
		sheet['H' + str(row_number)] = order["freight"]
		sheet['I' + str(row_number)] = order["advanced"]
		sheet['J' + str(row_number)] = float(order["freight"]) - float(order["advanced"])

		row_number += 1

	sheet["H19"] = str(total_freight)
	sheet["I19"] = str(total_advanced)
	sheet["J19"] = str(total_balance)
	sheet["J20"] = str(total_Liftoff)
	sheet["J21"] = str(total_balance + total_Liftoff)
	new_amount = (total_balance + total_Liftoff)//1
	sheet["B22"] = num2words(new_amount)


def style_sheet(sheet):

	font_name = "Arial"

	def set_font_height(font_name, size, row, column):
		font = Openpyxl_Style(name=font_name, size=size)
		sheet[column + row].font = font

	set_font_height(font_name, 28, str(1), "A")
	sheet.row_dimensions[1].height = 30

	set_font_height(font_name, 18, str(2), "A")
	sheet.row_dimensions[2].height = 20

	set_font_height(font_name, 14, str(3), "A")
	sheet.row_dimensions[3].height = 16

	for __i__ in range(6, 32):
		sheet.row_dimensions[__i__].height = 13

def create_bill_with_excel_template(orders, bill_number="", bill_date="", pan_number="", party_gst="", party_name=""):
    # file_path = "chat/mis_dashboard"
    try:

        work_book = Openpyxl_load_workbook(filename="files/template/Bill_V1.xlsx")
        sheet_abstract = work_book["abstract"]

        total_order = len(orders)

        for index in range(0,total_order, 10):
        	orders_temp = orders[index : index + 10]
        	new_sheet = work_book.copy_worksheet(sheet_abstract)
        	new_sheet.title = "Sheet" + str(index//10 + 1)

        	bill_number_tmp = bill_number + str(index//10 + 1)

        	new_sheet["H4"] = bill_number_tmp
        	new_sheet["H5"] = bill_date
        	new_sheet["H6"] = pan_number
        	new_sheet["B6"] = party_gst
        	new_sheet["B4"] = party_name

        	style_sheet(new_sheet)

        	load_data_into_sheet(orders_temp, new_sheet)

        work_book.remove_sheet(sheet_abstract)

        d = datetime.datetime.now()
        file_name = d.strftime("%d-%B-%Y_%I:%M_%p_%A") + "_" + party_name.strip()
        file_name = "files/exported_data/" + file_name + ".xlsx"
        work_book.save(file_name)
        # sheet_bkp = work_book.add_sheet("Bill1")
        # add_new_sheet_to_bill(sheet, orders)

        # sheet = work_book.add_sheet("Bill2")
        # add_new_sheet_to_bill(sheet, orders)

        # filename = "filtered_data_" + str(datetime.datetime.today())  + ".xls"

        # work_book.save(filename)

        # print("BHAI BHAI")
        # return filename
        return "http://127.0.0.1:8000/" + file_name
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        print("OH NO")
        print(str(e), str(exc_tb.tb_lineno))
